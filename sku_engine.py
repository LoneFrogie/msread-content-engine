"""
MS. READ Content Engine — SKU-Specific Content Generation
Fetches product data from msreadshop.com Shopify API, then generates:
  1. Blog post
  2. Social media posts (FB, IG, TikTok)
  3. Persuasive product descriptions
  4. Email campaign
"""

import json
import re
import time
from io import BytesIO
from pathlib import Path
from typing import Callable
from urllib.parse import urlparse
import zipfile
import logging

logger = logging.getLogger(__name__)

import requests
from bs4 import BeautifulSoup
from google import genai
from google.genai import types
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image as PILImage

# ── Brand constants ──
BRAND_NAME = "MS. READ"
BRAND_WEBSITE = "msreadshop.com"

# ── Excel colors (same as main engine) ──
NAVY = "1B2A4A"
GOLD = "C8A951"
BURGUNDY = "7A1F3D"
CREAM = "FFF8F0"
LIGHT_GOLD = "FFF3D6"
WHITE = "FFFFFF"
MED_GREY = "E0E0E0"
DARK_TEXT = "1A1A1A"

header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=NAVY)
body_font = Font(name="Arial", color=DARK_TEXT, size=10)
body_font_bold = Font(name="Arial", bold=True, color=DARK_TEXT, size=10)
accent_font = Font(name="Arial", bold=True, color=BURGUNDY, size=10)
title_font = Font(name="Arial", bold=True, color=NAVY, size=14)
subtitle_font = Font(name="Arial", bold=True, color=GOLD, size=11)
wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
center_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
thin_border = Border(
    left=Side(style="thin", color=MED_GREY),
    right=Side(style="thin", color=MED_GREY),
    top=Side(style="thin", color=MED_GREY),
    bottom=Side(style="thin", color=MED_GREY),
)
row1_fill = PatternFill("solid", fgColor=WHITE)
row2_fill = PatternFill("solid", fgColor=CREAM)

BRAND_IMAGE_PREFIX = (
    "Professional fashion editorial photography. "
    "Premium, inclusive, empowering brand aesthetic. "
    "Warm golden lighting, Malaysian setting. "
    "Plus-size model, confident and joyful. "
    "No text overlays. No watermarks. High resolution, photorealistic."
)


def _style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def _style_data_row(ws, row, cols, alt=False):
    fill = row2_fill if alt else row1_fill
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.fill = fill
        cell.alignment = wrap_align
        cell.border = thin_border


def _add_title_block(ws, title, subtitle, start_row=1):
    ws.cell(row=start_row, column=1, value=title).font = title_font
    ws.cell(row=start_row + 1, column=1, value=subtitle).font = subtitle_font
    return start_row + 3


# ═══════════════════════════════════════════════════════════════
# PHASE 1: Fetch product data from Shopify JSON API
# ═══════════════════════════════════════════════════════════════

def extract_handle_from_url(url: str) -> str:
    """Extract product handle from a msreadshop.com URL."""
    parsed = urlparse(url.strip())
    path = parsed.path.strip("/")

    # Handle: products/some-product-handle or collections/xxx/products/some-product-handle
    if "products/" in path:
        handle = path.split("products/")[-1].split("?")[0].split("#")[0]
        if handle:
            return handle

    # Maybe they pasted just the handle
    if "/" not in path and path:
        return path

    raise ValueError(f"Could not extract product handle from URL: {url}")


def fetch_product(url: str, callback: Callable) -> dict:
    """Fetch product data from msreadshop.com Shopify JSON API."""
    callback("status", {"phase": "fetching_product", "message": "Fetching product from msreadshop.com..."})

    handle = extract_handle_from_url(url)
    api_url = f"https://www.msreadshop.com/products/{handle}.json"

    resp = requests.get(api_url, timeout=15, headers={
        "User-Agent": "Mozilla/5.0 (compatible; MSReadContentEngine/1.0)"
    })

    if resp.status_code == 404:
        raise ValueError(f"Product not found: {handle}. Check the URL and try again.")
    resp.raise_for_status()

    product = resp.json().get("product", {})
    if not product:
        raise ValueError("Empty product data returned from API.")

    # Parse and structure the data
    parsed = _parse_product(product)
    callback("status", {
        "phase": "product_fetched",
        "message": f"Found: {parsed['title']} — RM{parsed['price']}",
        "product": {
            "title": parsed["title"],
            "price": parsed["price"],
            "compare_price": parsed["compare_price"],
            "image_url": parsed["image_url"],
            "product_type": parsed["product_type"],
            "sizes": parsed["sizes"],
            "colors": parsed["colors"],
        }
    })
    return parsed


def _parse_product(product: dict) -> dict:
    """Parse Shopify product JSON into a clean structure."""
    # Get price info from first available variant
    variants = product.get("variants", [])
    price = variants[0]["price"] if variants else "0.00"
    compare_price = variants[0].get("compare_at_price") if variants else None

    # Extract sizes and colors from options
    sizes = []
    colors = []
    for opt in product.get("options", []):
        if opt["name"].lower() == "size":
            sizes = opt.get("values", [])
        elif opt["name"].lower() == "color":
            colors = opt.get("values", [])

    # Parse body_html for product details
    body_html = product.get("body_html", "")
    details = _parse_body_html(body_html)

    # Get image URLs
    images = product.get("images", [])
    image_urls = [img["src"] for img in images]
    primary_image = images[0]["src"] if images else None

    return {
        "title": product.get("title", "Unknown Product"),
        "handle": product.get("handle", ""),
        "product_type": product.get("product_type", ""),
        "vendor": product.get("vendor", BRAND_NAME),
        "price": price,
        "compare_price": compare_price,
        "sizes": sizes,
        "colors": colors,
        "tags": product.get("tags", []),
        "image_url": primary_image,
        "image_urls": image_urls,
        "body_html": body_html,
        "description_text": details.get("description", ""),
        "fabric": details.get("fabric", ""),
        "care": details.get("care", ""),
        "sku": details.get("sku", variants[0].get("sku", "") if variants else ""),
        "url": f"https://www.msreadshop.com/products/{product.get('handle', '')}",
    }


def _parse_body_html(html: str) -> dict:
    """Extract structured info from Shopify body_html."""
    if not html:
        return {}

    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text(separator="\n", strip=True)

    result = {"description": text[:500]}

    # Look for common patterns in MS. READ product descriptions
    lines = text.split("\n")
    for line in lines:
        lower = line.lower().strip()
        if "sku" in lower and ":" in line:
            result["sku"] = line.split(":", 1)[1].strip()
        elif any(kw in lower for kw in ["polyester", "cotton", "linen", "nylon", "spandex", "viscose", "rayon"]):
            result["fabric"] = line.strip()
        elif any(kw in lower for kw in ["wash", "iron", "dry clean", "bleach"]):
            result["care"] = line.strip()

    return result


# ═══════════════════════════════════════════════════════════════
# PHASE 2: Generate content via Gemini text model
# ═══════════════════════════════════════════════════════════════

SKU_CONTENT_PROMPT = """You are a senior content strategist and copywriter for MS. READ, Malaysia's leading plus-size fashion brand (UK sizes 10-24+, founded 1997).

BRAND RULES (never break these):
- Voice: Warm, empowering, inclusive, confident
- NEVER use diet culture language or apologetic tone about size
- Target: Women 25-55, Malaysia & Singapore
- Always mention size range UK 10-24+
- CTA should direct to msreadshop.com or WhatsApp

PRODUCT DATA:
- Name: {title}
- Type: {product_type}
- Price: RM{price}{price_note}
- Sizes: {sizes}
- Colors: {colors}
- Fabric: {fabric}
- Description: {description}
- URL: {url}

{creative_brief_section}

Generate a comprehensive JSON object with these keys:

1. "blog_post": Object with:
   - "title": SEO-optimized blog title (include product name & keyword)
   - "meta_description": 155-char SEO meta description
   - "keyword": Primary target keyword
   - "introduction": 150-word opening paragraph (hook + problem)
   - "body_sections": Array of 3-4 section objects, each with "heading" and "content" (150-200 words each)
   - "conclusion": 100-word closing with CTA
   - "internal_links": Array of 3 suggested related product categories to link to

2. "social_posts": Array of 6 post objects, each with:
   - "platform": "Facebook" | "Instagram" | "TikTok" | "Instagram Stories" | "Instagram Reels" | "Facebook"
   - "format": "Carousel" | "Single Image" | "Reel/Short" | "Story" | "Post"
   - "caption": Full caption with emojis (platform-appropriate length)
   - "hashtags": 15-20 relevant hashtags as string
   - "hook": First line / attention grabber
   - "cta": Call to action
   - "notes": Content/creative direction notes for the post

3. "product_descriptions": Object with:
   - "short_description": 50-word punchy description for catalog/category pages
   - "long_description": 200-word detailed description for product page (persuasive, benefit-focused)
   - "whatsapp_description": 80-word casual description for WhatsApp broadcast
   - "marketplace_description": 150-word description for Shopee/Lazada (SEO keywords, bullet points)
   - "key_selling_points": Array of 5 benefit-focused selling points
   - "size_confidence": A reassuring size-inclusive statement for this specific product

4. "email_campaign": Object with:
   - "subject_lines": Array of 5 email subject lines (A/B test options)
   - "preview_text": Preview text that appears after subject line
   - "hero_headline": Bold headline for the email
   - "hero_subheadline": Supporting subheadline
   - "body_copy": 150-word persuasive body copy
   - "cta_button_text": CTA button text
   - "cta_url": Product URL
   - "urgency_line": Scarcity/urgency element
   - "ps_line": P.S. line at bottom of email
   - "segment_notes": Who this email should be sent to

5. "image_prompts": Array of 4 image prompt objects, each with:
   - "scene": Short label (e.g., "Lifestyle Shot", "Flat Lay", "Street Style", "Detail Close-up")
   - "prompt": Detailed image generation prompt (2-3 sentences, visually descriptive)
   - "usage": Where this image would be used (e.g., "Instagram carousel slide 1", "Blog hero image")

Use Malaysian English. Make all content specific to THIS product — not generic.
Respond with ONLY valid JSON, no markdown code fences."""


def generate_sku_content(client, product: dict, creative_brief: str, callback: Callable) -> dict:
    """Use Gemini to generate all SKU content."""
    callback("status", {"phase": "generating_content", "message": "Generating blog, social, descriptions, and email content..."})

    price_note = ""
    if product.get("compare_price"):
        price_note = f" (was RM{product['compare_price']})"

    creative_brief_section = ""
    if creative_brief.strip():
        creative_brief_section = f"USER'S CREATIVE DIRECTION:\n{creative_brief}\n\nAdapt all content to match this creative vision while keeping it specific to the product above."

    prompt = SKU_CONTENT_PROMPT.format(
        title=product["title"],
        product_type=product["product_type"],
        price=product["price"],
        price_note=price_note,
        sizes=", ".join(product["sizes"]) if product["sizes"] else "UK 10-24+",
        colors=", ".join(product["colors"]) if product["colors"] else "Multiple colors",
        fabric=product.get("fabric", "Premium fabric"),
        description=product.get("description_text", "")[:400],
        url=product["url"],
        creative_brief_section=creative_brief_section,
    )

    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=prompt,
        config=types.GenerateContentConfig(
            temperature=0.7,
            max_output_tokens=16000,
            response_mime_type="application/json",
        ),
    )

    text = ""
    for part in response.candidates[0].content.parts:
        if part.text:
            text += part.text

    text = text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1]
    if text.endswith("```"):
        text = text.rsplit("```", 1)[0]
    text = text.strip()

    content = _parse_json_robust(text)
    content = _sanitize_sku_content(content)
    callback("status", {"phase": "content_generated", "message": "All content generated successfully"})
    return content


def _parse_json_robust(text: str) -> dict:
    """Parse JSON with repair for common Gemini output issues."""
    # Try direct parse first
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Fix 1: Remove trailing commas before } or ]
    fixed = re.sub(r',\s*([\]}])', r'\1', text)
    try:
        return json.loads(fixed)
    except json.JSONDecodeError:
        pass

    # Fix 2: Escape unescaped newlines inside strings
    fixed2 = fixed.replace('\n', '\\n')
    try:
        return json.loads(fixed2)
    except json.JSONDecodeError:
        pass

    # Fix 3: Truncated JSON — try to close open brackets/braces
    repaired = fixed
    open_braces = repaired.count('{') - repaired.count('}')
    open_brackets = repaired.count('[') - repaired.count(']')

    # Trim to last complete value
    repaired = repaired.rstrip()
    if repaired and repaired[-1] not in '"}]0123456789truefalsenull':
        # Remove incomplete trailing content
        for end_char in ['"', '}', ']']:
            last_pos = repaired.rfind(end_char)
            if last_pos > 0:
                repaired = repaired[:last_pos + 1]
                break

    # Recount after trimming
    open_braces = repaired.count('{') - repaired.count('}')
    open_brackets = repaired.count('[') - repaired.count(']')
    repaired += ']' * max(0, open_brackets) + '}' * max(0, open_braces)

    try:
        return json.loads(repaired)
    except json.JSONDecodeError as e:
        logger.error(f"JSON repair failed. Raw text (first 500 chars): {text[:500]}")
        raise ValueError(f"Failed to parse Gemini response as JSON: {e}")


def _sanitize_sku_content(content: dict) -> dict:
    """Ensure all values are strings for Excel compatibility."""
    def to_str(val):
        if isinstance(val, list):
            return ", ".join(str(v) for v in val)
        if val is None:
            return ""
        return val

    # Sanitize blog
    blog = content.get("blog_post", {})
    for key in blog:
        if isinstance(blog[key], list):
            if key == "body_sections":
                for section in blog[key]:
                    for k in section:
                        section[k] = to_str(section[k])
            elif key == "internal_links":
                blog[key] = to_str(blog[key])
            else:
                blog[key] = to_str(blog[key])
        elif blog[key] is None:
            blog[key] = ""

    # Sanitize social posts
    for post in content.get("social_posts", []):
        for key in post:
            post[key] = to_str(post[key])

    # Sanitize product descriptions
    descs = content.get("product_descriptions", {})
    for key in descs:
        descs[key] = to_str(descs[key])

    # Sanitize email campaign
    email = content.get("email_campaign", {})
    for key in email:
        email[key] = to_str(email[key])

    # Sanitize image prompts
    for ip in content.get("image_prompts", []):
        for key in ip:
            ip[key] = to_str(ip[key])

    return content


# ═══════════════════════════════════════════════════════════════
# PHASE 3: Build Excel workbook
# ═══════════════════════════════════════════════════════════════

def build_sku_excel(content: dict, product: dict, output_dir: Path, callback: Callable) -> Path:
    """Build the 5-sheet Excel from SKU content."""
    callback("status", {"phase": "building_excel", "message": "Building Excel spreadsheet..."})

    wb = Workbook()
    product_title = product["title"]

    # ── SHEET 1: Product Overview ──
    ws1 = wb.active
    ws1.title = "Product Overview"
    ws1.sheet_properties.tabColor = NAVY

    r = _add_title_block(ws1, f"{BRAND_NAME}: SKU Content Pack", product_title)

    overview_data = [
        ["Product", product_title],
        ["Type", product["product_type"]],
        ["Price", f"RM{product['price']}"],
        ["Compare Price", f"RM{product['compare_price']}" if product.get("compare_price") else "—"],
        ["Sizes", ", ".join(product["sizes"]) if product["sizes"] else "UK 10-24+"],
        ["Colors", ", ".join(product["colors"]) if product["colors"] else "—"],
        ["Fabric", product.get("fabric", "—")],
        ["SKU", product.get("sku", "—")],
        ["URL", product["url"]],
    ]
    for item in overview_data:
        ws1.cell(row=r, column=1, value=item[0]).font = body_font_bold
        ws1.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_GOLD)
        ws1.cell(row=r, column=1).border = thin_border
        ws1.cell(row=r, column=2, value=item[1]).font = body_font
        ws1.cell(row=r, column=2).border = thin_border
        r += 1

    # Product descriptions section
    r += 1
    ws1.cell(row=r, column=1, value="PRODUCT DESCRIPTIONS").font = Font(name="Arial", bold=True, color=BURGUNDY, size=13)
    r += 1

    descs = content.get("product_descriptions", {})
    desc_rows = [
        ["Short Description", descs.get("short_description", "")],
        ["Long Description", descs.get("long_description", "")],
        ["WhatsApp Copy", descs.get("whatsapp_description", "")],
        ["Marketplace (Shopee/Lazada)", descs.get("marketplace_description", "")],
        ["Key Selling Points", descs.get("key_selling_points", "")],
        ["Size Confidence Statement", descs.get("size_confidence", "")],
    ]
    for item in desc_rows:
        ws1.cell(row=r, column=1, value=item[0]).font = body_font_bold
        ws1.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_GOLD)
        ws1.cell(row=r, column=1).border = thin_border
        ws1.cell(row=r, column=1).alignment = Alignment(vertical="top")
        ws1.cell(row=r, column=2, value=item[1]).font = body_font
        ws1.cell(row=r, column=2).alignment = wrap_align
        ws1.cell(row=r, column=2).border = thin_border
        ws1.row_dimensions[r].height = max(60, len(str(item[1])) // 3)
        r += 1

    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 100

    # ── SHEET 2: Blog Post ──
    ws2 = wb.create_sheet("Blog Post")
    ws2.sheet_properties.tabColor = "2D6A4F"

    blog = content.get("blog_post", {})
    r = _add_title_block(ws2, "SEO Blog Post", blog.get("title", product_title))

    meta = [
        ["Title", blog.get("title", "")],
        ["Meta Description", blog.get("meta_description", "")],
        ["Target Keyword", blog.get("keyword", "")],
        ["Internal Links", blog.get("internal_links", "")],
    ]
    for item in meta:
        ws2.cell(row=r, column=1, value=item[0]).font = body_font_bold
        ws2.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_GOLD)
        ws2.cell(row=r, column=1).border = thin_border
        ws2.cell(row=r, column=2, value=item[1]).font = body_font
        ws2.cell(row=r, column=2).alignment = wrap_align
        ws2.cell(row=r, column=2).border = thin_border
        r += 1

    r += 1
    # Introduction
    ws2.cell(row=r, column=1, value="INTRODUCTION").font = Font(name="Arial", bold=True, color=BURGUNDY, size=11)
    ws2.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_GOLD)
    ws2.cell(row=r, column=1).border = thin_border
    ws2.cell(row=r, column=2, value=blog.get("introduction", "")).font = body_font
    ws2.cell(row=r, column=2).alignment = wrap_align
    ws2.cell(row=r, column=2).border = thin_border
    ws2.row_dimensions[r].height = 100
    r += 1

    # Body sections
    for section in blog.get("body_sections", []):
        ws2.cell(row=r, column=1, value=section.get("heading", "")).font = Font(name="Arial", bold=True, color=BURGUNDY, size=11)
        ws2.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_GOLD)
        ws2.cell(row=r, column=1).border = thin_border
        ws2.cell(row=r, column=1).alignment = Alignment(vertical="top")
        ws2.cell(row=r, column=2, value=section.get("content", "")).font = body_font
        ws2.cell(row=r, column=2).alignment = wrap_align
        ws2.cell(row=r, column=2).border = thin_border
        ws2.row_dimensions[r].height = 120
        r += 1

    # Conclusion
    ws2.cell(row=r, column=1, value="CONCLUSION & CTA").font = Font(name="Arial", bold=True, color=BURGUNDY, size=11)
    ws2.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_GOLD)
    ws2.cell(row=r, column=1).border = thin_border
    ws2.cell(row=r, column=2, value=blog.get("conclusion", "")).font = body_font
    ws2.cell(row=r, column=2).alignment = wrap_align
    ws2.cell(row=r, column=2).border = thin_border
    ws2.row_dimensions[r].height = 80

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 100

    # ── SHEET 3: Social Media Posts ──
    ws3 = wb.create_sheet("Social Media Posts")
    ws3.sheet_properties.tabColor = GOLD

    r = _add_title_block(ws3, "Social Media Posts", f"{product_title} — Cross-Platform Content")

    social_headers = ["#", "Platform", "Format", "Hook", "Full Caption", "CTA", "Hashtags", "Notes"]
    for c, h in enumerate(social_headers, 1):
        ws3.cell(row=r, column=c, value=h)
    _style_header_row(ws3, r, len(social_headers))
    r += 1

    for i, post in enumerate(content.get("social_posts", [])):
        vals = [
            i + 1,
            post.get("platform", ""),
            post.get("format", ""),
            post.get("hook", ""),
            post.get("caption", ""),
            post.get("cta", ""),
            post.get("hashtags", ""),
            post.get("notes", ""),
        ]
        for c, v in enumerate(vals, 1):
            ws3.cell(row=r, column=c, value=v)
        _style_data_row(ws3, r, len(social_headers), alt=(i % 2 == 1))
        ws3.cell(row=r, column=1).font = body_font_bold
        ws3.cell(row=r, column=1).alignment = center_align
        ws3.row_dimensions[r].height = 180
        r += 1

    for col, w in zip("ABCDEFGH", [5, 16, 14, 35, 70, 30, 50, 35]):
        ws3.column_dimensions[col].width = w

    # ── SHEET 4: Email Campaign ──
    ws4 = wb.create_sheet("Email Campaign")
    ws4.sheet_properties.tabColor = "4A90D9"

    email = content.get("email_campaign", {})
    r = _add_title_block(ws4, "Email Campaign", f"{product_title} — Product Launch / Promotion Email")

    email_rows = [
        ["Subject Lines (A/B)", email.get("subject_lines", "")],
        ["Preview Text", email.get("preview_text", "")],
        ["Hero Headline", email.get("hero_headline", "")],
        ["Hero Subheadline", email.get("hero_subheadline", "")],
        ["Body Copy", email.get("body_copy", "")],
        ["CTA Button Text", email.get("cta_button_text", "")],
        ["CTA URL", email.get("cta_url", "")],
        ["Urgency Line", email.get("urgency_line", "")],
        ["P.S. Line", email.get("ps_line", "")],
        ["Target Segment", email.get("segment_notes", "")],
    ]
    for item in email_rows:
        ws4.cell(row=r, column=1, value=item[0]).font = body_font_bold
        ws4.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_GOLD)
        ws4.cell(row=r, column=1).border = thin_border
        ws4.cell(row=r, column=1).alignment = Alignment(vertical="top")
        ws4.cell(row=r, column=2, value=item[1]).font = body_font
        ws4.cell(row=r, column=2).alignment = wrap_align
        ws4.cell(row=r, column=2).border = thin_border
        ws4.row_dimensions[r].height = max(50, len(str(item[1])) // 3)
        r += 1

    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 100

    # ── SHEET 5: Image Prompts ──
    ws5 = wb.create_sheet("Image Prompts")
    ws5.sheet_properties.tabColor = "8B5E3C"

    r = _add_title_block(ws5, "Visual Direction & Image Prompts", f"{product_title} — AI Image Generation")

    img_headers = ["#", "Scene", "Detailed Prompt", "Usage"]
    for c, h in enumerate(img_headers, 1):
        ws5.cell(row=r, column=c, value=h)
    _style_header_row(ws5, r, len(img_headers))
    r += 1

    for i, ip in enumerate(content.get("image_prompts", [])):
        vals = [i + 1, ip.get("scene", ""), ip.get("prompt", ""), ip.get("usage", "")]
        for c, v in enumerate(vals, 1):
            ws5.cell(row=r, column=c, value=v)
        _style_data_row(ws5, r, len(img_headers), alt=(i % 2 == 1))
        ws5.cell(row=r, column=1).font = body_font_bold
        ws5.cell(row=r, column=1).alignment = center_align
        ws5.row_dimensions[r].height = 120
        r += 1

    for col, w in zip("ABCD", [5, 22, 85, 35]):
        ws5.column_dimensions[col].width = w

    # Save
    safe_title = re.sub(r'[^\w\s-]', '', product_title)[:50].strip().replace(' ', '_')
    excel_path = output_dir / f"MSRead_SKU_{safe_title}.xlsx"
    wb.save(excel_path)
    callback("status", {"phase": "excel_done", "message": f"Excel built (5 sheets) for {product_title}"})
    return excel_path


# ═══════════════════════════════════════════════════════════════
# PHASE 4: Generate AI images
# ═══════════════════════════════════════════════════════════════

def generate_sku_images(client, content: dict, product: dict, creative_brief: str,
                        output_dir: Path, callback: Callable) -> list:
    """Generate product images via Nano Banana."""
    image_dir = output_dir / "images"
    thumb_dir = output_dir / "thumbnails"
    image_dir.mkdir(parents=True, exist_ok=True)
    thumb_dir.mkdir(parents=True, exist_ok=True)

    prompts = content.get("image_prompts", [])
    total = len(prompts)
    callback("status", {"phase": "generating_images", "message": f"Generating {total} product images...", "total": total, "current": 0})

    product_context = f" Product: {product['title']}, {product['product_type']}."
    if creative_brief:
        product_context += f" Creative direction: {creative_brief[:150]}."

    generated_files = []

    for i, p in enumerate(prompts):
        scene = p.get("scene", f"scene_{i+1}")
        safe_scene = re.sub(r'[^\w]', '_', scene.lower())[:30]
        filename = f"sku_{safe_scene}.png"
        filepath = image_dir / filename

        callback("image_start", {
            "index": i, "total": total, "day": i + 1, "theme": scene,
            "message": f"Generating {scene}..."
        })

        full_prompt = f"{BRAND_IMAGE_PREFIX}{product_context}\n\n{p['prompt']}"

        try:
            response = client.models.generate_content(
                model="gemini-2.5-flash-image",
                contents=full_prompt,
                config=types.GenerateContentConfig(
                    response_modalities=["Text", "Image"],
                ),
            )

            image_saved = False
            for part in response.candidates[0].content.parts:
                if part.inline_data is not None:
                    img = PILImage.open(BytesIO(part.inline_data.data))
                    img.save(filepath, "PNG")

                    ratio = 200 / img.width
                    img_thumb = img.resize((200, int(img.height * ratio)), PILImage.LANCZOS)
                    img_thumb.save(thumb_dir / filename, "PNG")
                    image_saved = True
                    break

            if image_saved:
                generated_files.append({"scene": scene, "filename": filename})
                callback("image_done", {
                    "index": i, "total": total, "day": i + 1, "theme": scene,
                    "filename": filename, "success": True,
                    "message": f"{scene} done"
                })
            else:
                callback("image_done", {
                    "index": i, "total": total, "day": i + 1, "theme": scene,
                    "filename": None, "success": False,
                    "message": f"{scene} — No image returned"
                })

        except Exception as e:
            callback("image_done", {
                "index": i, "total": total, "day": i + 1, "theme": scene,
                "filename": None, "success": False,
                "message": f"{scene} — Failed: {str(e)[:100]}"
            })

        if i < total - 1:
            time.sleep(8)

    return generated_files


# ═══════════════════════════════════════════════════════════════
# PHASE 5: Package everything
# ═══════════════════════════════════════════════════════════════

def package_sku_output(output_dir: Path, product_title: str, callback: Callable) -> Path:
    """Create zip of Excel + images."""
    callback("status", {"phase": "packaging", "message": "Creating download package..."})

    safe_title = re.sub(r'[^\w\s-]', '', product_title)[:50].strip().replace(' ', '_')
    zip_path = output_dir / f"MSRead_SKU_{safe_title}.zip"

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        # Add Excel
        for xlsx in output_dir.glob("*.xlsx"):
            zf.write(xlsx, xlsx.name)
        # Add images
        image_dir = output_dir / "images"
        if image_dir.exists():
            for img_file in sorted(image_dir.glob("*.png")):
                zf.write(img_file, f"images/{img_file.name}")

    size_mb = zip_path.stat().st_size / (1024 * 1024)
    callback("status", {
        "phase": "done",
        "message": f"Done! Package ready ({size_mb:.1f} MB)",
        "zip_size_mb": round(size_mb, 1),
    })
    return zip_path


# ═══════════════════════════════════════════════════════════════
# ORCHESTRATOR
# ═══════════════════════════════════════════════════════════════

def run_sku_pipeline(api_key: str, product_url: str, creative_brief: str,
                     output_dir: Path, callback: Callable):
    """Run the full SKU content pipeline."""
    try:
        client = genai.Client(api_key=api_key)

        # Phase 1: Fetch product
        product = fetch_product(product_url, callback)

        # Phase 2: Generate content
        content = generate_sku_content(client, product, creative_brief, callback)

        # Phase 3: Build Excel
        build_sku_excel(content, product, output_dir, callback)

        # Phase 4: Generate images
        generate_sku_images(client, content, product, creative_brief, output_dir, callback)

        # Phase 5: Package
        package_sku_output(output_dir, product["title"], callback)

    except Exception as e:
        callback("error", {"message": f"Pipeline failed: {str(e)}"})
        raise
