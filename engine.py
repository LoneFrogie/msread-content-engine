"""
MS. READ Content Engine — Core Generation Logic
Takes a creative brief and produces a 30-day content calendar Excel + AI images.
"""

import json
import time
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta
from typing import Callable, Optional
import zipfile

from google import genai
from google.genai import types
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# ── Brand constants ──
BRAND_NAME = "MS. READ"
BRAND_WEBSITE = "msreadshop.com"
BRAND_VOICE = "Warm, empowering, inclusive, confident"

# ── Excel color constants ──
NAVY = "1B2A4A"
GOLD = "C8A951"
BURGUNDY = "7A1F3D"
CREAM = "FFF8F0"
LIGHT_GOLD = "FFF3D6"
WHITE = "FFFFFF"
MED_GREY = "E0E0E0"
DARK_TEXT = "1A1A1A"

# ── Excel style helpers ──
header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=NAVY)
body_font = Font(name="Arial", color=DARK_TEXT, size=10)
body_font_bold = Font(name="Arial", bold=True, color=DARK_TEXT, size=10)
accent_font = Font(name="Arial", bold=True, color=BURGUNDY, size=10)
title_font = Font(name="Arial", bold=True, color=NAVY, size=14)
subtitle_font = Font(name="Arial", bold=True, color=GOLD, size=11)
wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
center_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
thin_border = Border(
    left=Side(style="thin", color=MED_GREY),
    right=Side(style="thin", color=MED_GREY),
    top=Side(style="thin", color=MED_GREY),
    bottom=Side(style="thin", color=MED_GREY),
)
row1_fill = PatternFill("solid", fgColor=WHITE)
row2_fill = PatternFill("solid", fgColor=CREAM)

BRAND_IMAGE_PREFIX = (
    "Professional fashion editorial photography. "
    "Premium, inclusive, empowering brand aesthetic. "
    "Warm golden lighting, Malaysian setting. "
    "Plus-size model, confident and joyful. "
    "No text overlays. No watermarks. High resolution, photorealistic."
)


def _style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def _style_data_row(ws, row, cols, alt=False):
    fill = row2_fill if alt else row1_fill
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.fill = fill
        cell.alignment = wrap_align
        cell.border = thin_border


def _add_title_block(ws, title, subtitle, start_row=1):
    ws.cell(row=start_row, column=1, value=title).font = title_font
    ws.cell(row=start_row + 1, column=1, value=subtitle).font = subtitle_font
    return start_row + 3


# ═══════════════════════════════════════════════════════════════
# PHASE 1: Generate adapted content via Gemini text model
# ═══════════════════════════════════════════════════════════════

BRAND_RULES = """BRAND RULES (never break these):
- Voice: Warm, empowering, inclusive, confident
- NEVER use diet culture language or apologetic tone about size
- Target: Women 25-55, Malaysia & Singapore
- Always mention size range UK 10-24+
- CTA should direct to msreadshop.com or WhatsApp
- Use Malaysian English nuances"""

# ── Call 1: Calendar + Trends + Hashtag Bank ──
PROMPT_CALENDAR = """You are a social media strategist for MS. READ, Malaysia's leading plus-size fashion brand (UK sizes 10-24+, founded 1997).

{brand_rules}

CREATIVE DIRECTION FROM USER:
{creative_brief}

Generate a JSON object with these keys:

1. "title_subtitle": Brief subtitle for the calendar (e.g., "Raya Collection Focus — March 2026")

2. "hashtag_bank": Object with:
   - "core": Array of 5-7 hashtags that MUST appear on EVERY single post (brand identity hashtags like #MSRead #FashionThatFits etc.)
   - "campaign": Array of 5-8 campaign-specific hashtags for this creative direction
   - "engagement": Array of 5-8 engagement-driving hashtags (#OOTD, #StyleInspo etc.)
   - "local": Array of 5-8 Malaysian/local hashtags

3. "trends": Array of 3 trend objects, each with: "name", "platform", "audio", "description", "adaptation", "engagement", "frequency"

4. "calendar": Array of 30 day objects, each with: "day" (1-30), "weekday", "date", "platforms", "format", "content_type", "hook", "trend_connection", "value_body", "cta", "audio", "hashtags"
   IMPORTANT: Every day's "hashtags" MUST include ALL core hashtags from the hashtag_bank, plus a selection of campaign/engagement/local hashtags.

5. "image_prompts": Array of 30 image prompt objects (ONE per day), each with: "day" (1-30), "date", "theme", "prompt" (detailed image generation prompt, 2-3 sentences, visually descriptive), "key_elements", "colors"

Start dates from {start_date}. Make hooks attention-grabbing and specific to the creative direction.

Respond with ONLY valid JSON, no markdown code fences."""

# ── Call 2: Blog Posts (4) ──
PROMPT_BLOGS = """You are a senior content writer for MS. READ, Malaysia's leading plus-size fashion brand (UK sizes 10-24+, founded 1997).

{brand_rules}

CREATIVE DIRECTION:
{creative_brief}

CORE HASHTAGS (include in all content):
{core_hashtags}

Generate a JSON object with ONE key:

"blogs": Array of 4 blog post objects. Each blog should cover a DIFFERENT angle of the creative direction. Each object has:
  - "title": SEO-optimized blog title
  - "keyword": Primary target keyword
  - "meta_description": 155-character SEO meta description
  - "problem": The problem/pain point (150 words)
  - "agitate": Agitate the problem (150 words)
  - "solution": MS. READ's solution (200 words)
  - "action_steps": Practical tips/steps (200 words)
  - "cta": Call to action (50 words)
  - "internal_links": Array of 3 suggested product category links

Blog topics should be:
1. Main campaign theme overview
2. Styling guide / how-to
3. Occasion-specific guide (e.g., office, events, casual)
4. Confidence / empowerment angle

Respond with ONLY valid JSON."""

# ── Call 3: Social Captions (40) ──
PROMPT_CAPTIONS = """You are a social media copywriter for MS. READ, Malaysia's leading plus-size fashion brand (UK sizes 10-24+, founded 1997).

{brand_rules}

CREATIVE DIRECTION:
{creative_brief}

CORE HASHTAGS (MUST appear on EVERY caption):
{core_hashtags}

CAMPAIGN HASHTAGS (use at least 3 per caption):
{campaign_hashtags}

Generate a JSON object with ONE key:

"social_copy": Array of 40 Facebook/Meta caption objects. Cover the full 30-day period with varied themes. Each object has:
  - "post_num": Sequential number (1-40)
  - "date": Date from the calendar period starting {start_date}
  - "theme": Post theme/topic
  - "caption": Full caption with emojis (150-300 words, engaging, Malaysian English)
  - "hashtags": String of 15-20 hashtags (MUST include ALL core hashtags + at least 3 campaign hashtags + relevant engagement/local hashtags)

Vary the content types across:
- Product spotlights (10 posts)
- Styling tips & outfit ideas (8 posts)
- Behind-the-scenes & brand story (4 posts)
- Customer testimonials / UGC prompts (5 posts)
- Trend-based / timely content (5 posts)
- Engagement / polls / questions (4 posts)
- Sales / promotions / new arrivals (4 posts)

Respond with ONLY valid JSON."""


def _gemini_call(client, prompt: str, max_tokens: int = 32000, temperature: float = 0.7) -> dict:
    """Make a Gemini text call with retry on JSON parse failure."""
    import re as _re
    last_error = None
    for attempt, temp in enumerate([temperature, 0.4]):
        try:
            response = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=prompt,
                config=types.GenerateContentConfig(
                    temperature=temp,
                    max_output_tokens=max_tokens,
                    response_mime_type="application/json",
                ),
            )
            text = ""
            for part in response.candidates[0].content.parts:
                if part.text:
                    text += part.text
            text = text.strip()
            if text.startswith("```"):
                text = text.split("\n", 1)[1]
            if text.endswith("```"):
                text = text.rsplit("```", 1)[0]
            text = text.strip()

            # Try direct parse
            try:
                return json.loads(text)
            except json.JSONDecodeError:
                pass
            # Fix trailing commas
            fixed = _re.sub(r',\s*([\]}])', r'\1', text)
            try:
                return json.loads(fixed)
            except json.JSONDecodeError:
                pass
            # Close truncated JSON
            repaired = fixed.rstrip()
            open_b = repaired.count('{') - repaired.count('}')
            open_k = repaired.count('[') - repaired.count(']')
            repaired += ']' * max(0, open_k) + '}' * max(0, open_b)
            return json.loads(repaired)
        except (json.JSONDecodeError, ValueError) as e:
            last_error = e
            if attempt == 0:
                time.sleep(2)
    raise last_error


def generate_adapted_content(client, creative_brief: str, start_date: str, callback: Callable):
    """Use 3 Gemini calls to generate the full content pack."""

    # ── Call 1: Calendar + Trends + Hashtag Bank + 30 Image Prompts ──
    callback("status", {"phase": "adapting", "message": "Generating 30-day calendar, trends, hashtags & image prompts..."})
    cal_prompt = PROMPT_CALENDAR.format(
        brand_rules=BRAND_RULES,
        creative_brief=creative_brief,
        start_date=start_date,
    )
    content = _gemini_call(client, cal_prompt, max_tokens=60000)
    content = _sanitize_content(content)
    callback("status", {"phase": "adapting", "message": "Calendar done. Generating 4 blog posts..."})

    # Extract hashtag bank for consistency
    hashtag_bank = content.get("hashtag_bank", {})
    core_hashtags = ", ".join(hashtag_bank.get("core", ["#MSRead", "#FashionThatFits", "#PlusSizeFashion"]))
    campaign_hashtags = ", ".join(hashtag_bank.get("campaign", []))

    # ── Call 2: 4 Blog Posts ──
    blog_prompt = PROMPT_BLOGS.format(
        brand_rules=BRAND_RULES,
        creative_brief=creative_brief,
        core_hashtags=core_hashtags,
    )
    blog_result = _gemini_call(client, blog_prompt, max_tokens=32000)
    content["blogs"] = blog_result.get("blogs", [])
    # Keep backward compat: also store first blog as "blog"
    if content["blogs"]:
        content["blog"] = content["blogs"][0]
    callback("status", {"phase": "adapting", "message": "Blogs done. Generating 40 social captions..."})

    # ── Call 3: 40 Social Captions ──
    caption_prompt = PROMPT_CAPTIONS.format(
        brand_rules=BRAND_RULES,
        creative_brief=creative_brief,
        core_hashtags=core_hashtags,
        campaign_hashtags=campaign_hashtags,
        start_date=start_date,
    )
    caption_result = _gemini_call(client, caption_prompt, max_tokens=60000)
    content["social_copy"] = caption_result.get("social_copy", [])

    # Sanitize new content
    for blog in content.get("blogs", []):
        for key in blog:
            if isinstance(blog[key], list):
                blog[key] = ", ".join(str(v) for v in blog[key])
            elif blog[key] is None:
                blog[key] = ""
    for cap in content.get("social_copy", []):
        for key in ["theme", "caption", "hashtags", "date"]:
            if key in cap:
                val = cap[key]
                if isinstance(val, list):
                    cap[key] = ", ".join(str(v) for v in val)
                elif val is None:
                    cap[key] = ""

    callback("status", {"phase": "adapted", "message": "All content generated: 30-day calendar, 4 blogs, 40 captions, 30 image prompts"})
    return content


def _sanitize_content(content: dict) -> dict:
    """Ensure all values that go into Excel cells are strings, not lists."""
    def to_str(val):
        if isinstance(val, list):
            return ", ".join(str(v) for v in val)
        if val is None:
            return ""
        return val

    # Sanitize calendar entries
    for day in content.get("calendar", []):
        for key in ["platforms", "format", "content_type", "hook", "trend_connection",
                     "value_body", "cta", "audio", "hashtags", "weekday", "date"]:
            if key in day:
                day[key] = to_str(day[key])

    # Sanitize trends
    for t in content.get("trends", []):
        for key in t:
            t[key] = to_str(t[key])

    # Sanitize social copy
    for cap in content.get("social_copy", []):
        for key in ["theme", "caption", "hashtags", "date"]:
            if key in cap:
                cap[key] = to_str(cap[key])

    # Sanitize image prompts
    for ip in content.get("image_prompts", []):
        for key in ["theme", "prompt", "key_elements", "colors", "date"]:
            if key in ip:
                ip[key] = to_str(ip[key])

    # Sanitize blog
    blog = content.get("blog", {})
    for key in blog:
        blog[key] = to_str(blog[key])

    return content


# ═══════════════════════════════════════════════════════════════
# PHASE 2: Build Excel from adapted content
# ═══════════════════════════════════════════════════════════════

def build_excel(content: dict, output_dir: Path, callback: Callable) -> Path:
    """Build the 6-sheet Excel from adapted content data."""
    callback("status", {"phase": "building_excel", "message": "Building Excel spreadsheet..."})

    wb = Workbook()
    start_date = content.get("calendar", [{}])[0].get("date", "")
    subtitle = content.get("title_subtitle", f"30-Day Content Engine — {start_date}")

    # ── SHEET 1: Overview & Trends ──
    ws1 = wb.active
    ws1.title = "Overview & Trends"
    ws1.sheet_properties.tabColor = NAVY

    r = _add_title_block(ws1, f"{BRAND_NAME}: 30-Day Viral-to-Value Content Engine", subtitle)

    overview = [
        ["Brand", BRAND_NAME], ["Website", BRAND_WEBSITE],
        ["Founded", "1997 by Helen Read"],
        ["Target Audience", "Women 25-55, UK 10-24+"],
        ["Markets", "Malaysia & Singapore"],
        ["Brand Voice", BRAND_VOICE],
        ["Creative Direction", content.get("title_subtitle", "")],
    ]
    for item in overview:
        ws1.cell(row=r, column=1, value=item[0]).font = body_font_bold
        ws1.cell(row=r, column=2, value=item[1]).font = body_font
        r += 1

    r += 1
    ws1.cell(row=r, column=1, value="TREND MAPPING").font = Font(name="Arial", bold=True, color=BURGUNDY, size=13)
    r += 1

    trend_headers = ["Trend", "Platform", "Audio", "What It Is", "MS. READ Adaptation", "Engagement Potential", "Post Frequency"]
    for c, h in enumerate(trend_headers, 1):
        ws1.cell(row=r, column=c, value=h)
    _style_header_row(ws1, r, len(trend_headers))
    r += 1

    for i, t in enumerate(content.get("trends", [])):
        vals = [t.get("name", ""), t.get("platform", ""), t.get("audio", ""),
                t.get("description", ""), t.get("adaptation", ""),
                t.get("engagement", ""), t.get("frequency", "")]
        for c, v in enumerate(vals, 1):
            ws1.cell(row=r, column=c, value=v)
        _style_data_row(ws1, r, len(trend_headers), alt=(i % 2 == 1))
        ws1.row_dimensions[r].height = 80
        r += 1

    for col, w in zip("ABCDEFG", [28, 18, 30, 50, 55, 40, 14]):
        ws1.column_dimensions[col].width = w

    # ── SHEET 2: 30-Day Calendar ──
    ws2 = wb.create_sheet("30-Day Calendar")
    ws2.sheet_properties.tabColor = BURGUNDY

    r = _add_title_block(ws2, "30-Day Content Calendar", subtitle)

    cal_headers = ["Day #", "Day", "Date", "Platform(s)", "Format", "Content Type",
                   "Hook (0-5 seconds)", "Trend Connection", "Value / Body", "CTA", "Audio / Sound", "Hashtag Set"]
    for c, h in enumerate(cal_headers, 1):
        ws2.cell(row=r, column=c, value=h)
    _style_header_row(ws2, r, len(cal_headers))
    r += 1

    for i, day in enumerate(content.get("calendar", [])):
        vals = [day.get("day", i+1), day.get("weekday", ""), day.get("date", ""),
                day.get("platforms", ""), day.get("format", ""), day.get("content_type", ""),
                day.get("hook", ""), day.get("trend_connection", ""),
                day.get("value_body", ""), day.get("cta", ""),
                day.get("audio", ""), day.get("hashtags", "")]
        for c, v in enumerate(vals, 1):
            ws2.cell(row=r, column=c, value=v)
        _style_data_row(ws2, r, len(cal_headers), alt=(i % 2 == 1))
        ws2.row_dimensions[r].height = 80
        r += 1

    for col, w in zip("ABCDEFGHIJKL", [6, 6, 10, 18, 22, 14, 40, 30, 45, 35, 25, 40]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A4"

    # ── SHEET 3: Blog Posts (4) ──
    ws3 = wb.create_sheet("Blog Posts")
    ws3.sheet_properties.tabColor = "2D6A4F"

    blogs = content.get("blogs", [])
    if not blogs and content.get("blog"):
        blogs = [content["blog"]]

    r = _add_title_block(ws3, "SEO Blog Posts", f"{len(blogs)} Blog Posts — Problem-Agitate-Solution (PAS) Framework")

    for blog_idx, blog in enumerate(blogs):
        # Blog header
        ws3.cell(row=r, column=1, value=f"BLOG {blog_idx + 1}").font = Font(name="Arial", bold=True, color=WHITE, size=12)
        ws3.cell(row=r, column=1).fill = PatternFill("solid", fgColor=BURGUNDY)
        ws3.cell(row=r, column=2, value=blog.get("title", "")).font = Font(name="Arial", bold=True, color=WHITE, size=12)
        ws3.cell(row=r, column=2).fill = PatternFill("solid", fgColor=BURGUNDY)
        ws3.cell(row=r, column=1).border = thin_border
        ws3.cell(row=r, column=2).border = thin_border
        r += 1

        meta = [
            ["Title", blog.get("title", "")],
            ["Keyword", blog.get("keyword", "")],
            ["Meta Description", blog.get("meta_description", "")],
            ["Internal Links", blog.get("internal_links", "")],
        ]
        for item in meta:
            ws3.cell(row=r, column=1, value=item[0]).font = body_font_bold
            ws3.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_GOLD)
            ws3.cell(row=r, column=2, value=item[1]).font = body_font
            ws3.cell(row=r, column=1).border = thin_border
            ws3.cell(row=r, column=2).border = thin_border
            ws3.cell(row=r, column=2).alignment = wrap_align
            r += 1

        sections = [
            ["PROBLEM", blog.get("problem", "")],
            ["AGITATE", blog.get("agitate", "")],
            ["SOLUTION", blog.get("solution", "")],
            ["ACTION STEPS", blog.get("action_steps", "")],
            ["CTA", blog.get("cta", "")],
        ]
        for section in sections:
            ws3.cell(row=r, column=1, value=section[0]).font = Font(name="Arial", bold=True, color=BURGUNDY, size=11)
            ws3.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_GOLD)
            ws3.cell(row=r, column=1).border = thin_border
            ws3.cell(row=r, column=1).alignment = Alignment(vertical="top")
            ws3.cell(row=r, column=2, value=section[1]).font = body_font
            ws3.cell(row=r, column=2).alignment = wrap_align
            ws3.cell(row=r, column=2).border = thin_border
            ws3.row_dimensions[r].height = max(80, len(str(section[1])) // 3)
            r += 1

        r += 1  # spacing between blogs

    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 100

    # ── SHEET 4: Social Copy (40 captions) ──
    ws4 = wb.create_sheet("Social Copy")
    ws4.sheet_properties.tabColor = GOLD

    social_copy = content.get("social_copy", [])
    r = _add_title_block(ws4, "Facebook / Meta Captions", f"{len(social_copy)} Captions — 30-Day Campaign")

    copy_headers = ["Post #", "Date", "Post Theme", "Full Caption", "Hashtags"]
    for c, h in enumerate(copy_headers, 1):
        ws4.cell(row=r, column=c, value=h)
    _style_header_row(ws4, r, len(copy_headers))
    r += 1

    for i, cap in enumerate(social_copy):
        vals = [cap.get("post_num", i+1), cap.get("date", ""), cap.get("theme", ""),
                cap.get("caption", ""), cap.get("hashtags", "")]
        for c, v in enumerate(vals, 1):
            ws4.cell(row=r, column=c, value=v)
        _style_data_row(ws4, r, len(copy_headers), alt=(i % 2 == 1))
        ws4.cell(row=r, column=1).font = body_font_bold
        ws4.cell(row=r, column=1).alignment = center_align
        ws4.row_dimensions[r].height = 200
        r += 1

    for col, w in zip("ABCDE", [8, 12, 28, 80, 50]):
        ws4.column_dimensions[col].width = w

    # ── SHEET 5: Image Prompts ──
    ws5 = wb.create_sheet("Image Prompts")
    ws5.sheet_properties.tabColor = "8B5E3C"

    r = _add_title_block(ws5, "Visual Direction & Image Generation Prompts", "Nano Banana (Gemini) — Premium, Inclusive, Malaysian Aesthetic")

    guidelines = [
        ["Color Palette", "Warm neutrals (cream, camel, soft gold), accents of deep burgundy and forest green"],
        ["Lighting", "Golden hour / warm studio lighting. Never harsh or clinical"],
        ["Models", "Diverse body types (UK 14-24), diverse ethnicities (Malay, Chinese, Indian, mixed)"],
        ["Setting", "Premium but accessible — upscale Malaysian locations, not untouchable luxury"],
        ["Mood", "Confident, joyful, aspirational, inclusive"],
    ]
    ws5.cell(row=r, column=1, value="BRAND AESTHETIC GUIDELINES").font = accent_font
    r += 1
    for g in guidelines:
        ws5.cell(row=r, column=1, value=g[0]).font = body_font_bold
        ws5.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_GOLD)
        ws5.cell(row=r, column=2, value=g[1]).font = body_font
        ws5.cell(row=r, column=1).border = thin_border
        ws5.cell(row=r, column=2).border = thin_border
        r += 1

    r += 1
    img_headers = ["Day #", "Date", "Post Theme", "Detailed Image Generation Prompt", "Key Visual Elements", "Color Palette"]
    for c, h in enumerate(img_headers, 1):
        ws5.cell(row=r, column=c, value=h)
    _style_header_row(ws5, r, len(img_headers))
    r += 1

    for i, ip in enumerate(content.get("image_prompts", [])):
        vals = [ip.get("day", ""), ip.get("date", ""), ip.get("theme", ""),
                ip.get("prompt", ""), ip.get("key_elements", ""), ip.get("colors", "")]
        for c, v in enumerate(vals, 1):
            ws5.cell(row=r, column=c, value=v)
        _style_data_row(ws5, r, len(img_headers), alt=(i % 2 == 1))
        ws5.cell(row=r, column=1).font = body_font_bold
        ws5.cell(row=r, column=1).alignment = center_align
        ws5.row_dimensions[r].height = 120
        r += 1

    for col, w in zip("ABCDEF", [8, 10, 22, 85, 45, 30]):
        ws5.column_dimensions[col].width = w

    # ── SHEET 6: Hashtags & Schedule ──
    ws6 = wb.create_sheet("Hashtags & Schedule")
    ws6.sheet_properties.tabColor = "4A90D9"

    r = _add_title_block(ws6, "Hashtag Banks & Posting Schedule", "Optimized for GMT+8 (Malaysia/Singapore)")

    ws6.cell(row=r, column=1, value="HASHTAG BANKS").font = accent_font
    r += 1
    hashtag_bank = content.get("hashtag_bank", {})
    hashtag_sets = [
        ["Core (EVERY Post)", " ".join(hashtag_bank.get("core", ["#MSRead", "#FashionThatFits", "#PlusSizeFashion"]))],
        ["Campaign-Specific", " ".join(hashtag_bank.get("campaign", []))],
        ["Engagement-Driving", " ".join(hashtag_bank.get("engagement", ["#OOTD", "#StyleInspo"]))],
        ["Malaysian / Local", " ".join(hashtag_bank.get("local", ["#PlusSizeMalaysia", "#MalaysianFashion"]))],
    ]
    for hs in hashtag_sets:
        ws6.cell(row=r, column=1, value=hs[0]).font = body_font_bold
        ws6.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_GOLD)
        ws6.cell(row=r, column=2, value=hs[1]).font = body_font
        ws6.cell(row=r, column=1).border = thin_border
        ws6.cell(row=r, column=2).border = thin_border
        r += 1

    r += 2
    ws6.cell(row=r, column=1, value="OPTIMAL POSTING SCHEDULE (GMT+8)").font = accent_font
    r += 1
    sched_headers = ["Platform", "Best Days", "Best Times", "Content Type"]
    for c, h in enumerate(sched_headers, 1):
        ws6.cell(row=r, column=c, value=h)
    _style_header_row(ws6, r, len(sched_headers))
    r += 1
    schedule = [
        ["TikTok", "Mon, Wed, Fri, Sun", "12pm-1pm, 7pm-9pm", "Reels, GRWM, Trends"],
        ["Instagram Feed", "Tue, Thu, Sat", "11am-1pm, 7pm-8pm", "Carousels, UGC, Promos"],
        ["Instagram Reels", "Mon, Wed, Fri", "12pm-2pm, 8pm-9pm", "Trend adaptations"],
        ["Facebook", "Tue, Thu, Sat", "10am-12pm, 3pm-4pm", "Longer captions, community"],
    ]
    for i, row_data in enumerate(schedule):
        for c, v in enumerate(row_data, 1):
            ws6.cell(row=r, column=c, value=v)
        _style_data_row(ws6, r, len(sched_headers), alt=(i % 2 == 1))
        r += 1

    for col, w in zip("ABCD", [32, 55, 35, 30]):
        ws6.column_dimensions[col].width = w

    # Save
    excel_path = output_dir / "MSRead_Content_Engine.xlsx"
    wb.save(excel_path)
    callback("status", {"phase": "excel_done", "message": "Excel built (6 sheets, 30 days)"})
    return excel_path


# ═══════════════════════════════════════════════════════════════
# PHASE 3: Generate AI images
# ═══════════════════════════════════════════════════════════════

def extract_prompts_from_excel(excel_path: Path) -> list[dict]:
    """Read image prompts from the Image Prompts sheet."""
    wb = load_workbook(excel_path, read_only=True)
    ws = wb["Image Prompts"]

    prompts = []
    header_found = False
    for row in ws.iter_rows(min_row=1, values_only=False):
        values = [cell.value for cell in row]
        if not header_found:
            if values and values[0] == "Day #":
                header_found = True
            continue
        if not values[0]:
            continue
        if values[3]:
            prompts.append({
                "day": values[0],
                "date": str(values[1] or ""),
                "theme": str(values[2] or ""),
                "prompt": str(values[3]),
            })
    wb.close()
    return prompts


def generate_images(client, creative_brief: str, excel_path: Path, output_dir: Path, callback: Callable):
    """Generate all images from Excel prompts via Nano Banana."""
    image_dir = output_dir / "images"
    thumb_dir = output_dir / "thumbnails"
    image_dir.mkdir(parents=True, exist_ok=True)
    thumb_dir.mkdir(parents=True, exist_ok=True)

    prompts = extract_prompts_from_excel(excel_path)
    total = len(prompts)
    callback("status", {"phase": "generating_images", "message": f"Generating {total} images...", "total": total, "current": 0})

    brief_context = ""
    if creative_brief:
        brief_context = f" Creative direction: {creative_brief[:200]}."

    generated_files = []

    for i, p in enumerate(prompts):
        day = p["day"]
        theme = p["theme"]
        safe_theme = theme.lower().replace(" ", "_").replace("/", "_").replace("'", "")
        filename = f"day{int(day):02d}_{safe_theme}.png"
        filepath = image_dir / filename

        callback("image_start", {
            "index": i, "total": total, "day": day, "theme": theme,
            "message": f"Generating Day {day} — {theme}..."
        })

        full_prompt = f"{BRAND_IMAGE_PREFIX}{brief_context}\n\n{p['prompt']}"

        try:
            response = client.models.generate_content(
                model="gemini-2.5-flash-image",
                contents=full_prompt,
                config=types.GenerateContentConfig(
                    response_modalities=["Text", "Image"],
                ),
            )

            image_saved = False
            for part in response.candidates[0].content.parts:
                if part.inline_data is not None:
                    img = PILImage.open(BytesIO(part.inline_data.data))
                    img.save(filepath, "PNG")

                    # Thumbnail
                    ratio = 200 / img.width
                    img_thumb = img.resize((200, int(img.height * ratio)), PILImage.LANCZOS)
                    img_thumb.save(thumb_dir / filename, "PNG")
                    image_saved = True
                    break

            if image_saved:
                generated_files.append({"day": day, "theme": theme, "filename": filename})
                callback("image_done", {
                    "index": i, "total": total, "day": day, "theme": theme,
                    "filename": filename, "success": True,
                    "message": f"Day {day} — {theme} done"
                })
            else:
                callback("image_done", {
                    "index": i, "total": total, "day": day, "theme": theme,
                    "filename": None, "success": False,
                    "message": f"Day {day} — No image returned"
                })

        except Exception as e:
            callback("image_done", {
                "index": i, "total": total, "day": day, "theme": theme,
                "filename": None, "success": False,
                "message": f"Day {day} — Failed: {str(e)[:100]}"
            })

        # Rate limiting
        if i < total - 1:
            time.sleep(8)

    # Embed thumbnails into Excel
    callback("status", {"phase": "embedding", "message": "Embedding thumbnails into Excel..."})
    _embed_thumbnails(excel_path, thumb_dir)

    return generated_files


def _embed_thumbnails(excel_path: Path, thumb_dir: Path):
    """Embed thumbnail images into the Image Prompts sheet."""
    wb = load_workbook(excel_path)
    ws = wb["Image Prompts"]

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=30):
        for cell in row:
            if cell.value == "Day #":
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        wb.close()
        return

    # Add Preview column
    max_col = 0
    for cell in ws[header_row]:
        if cell.value is not None:
            max_col = cell.column
    preview_col = max_col + 1
    ws.cell(row=header_row, column=preview_col, value="Preview")
    ws.cell(row=header_row, column=preview_col).font = header_font
    ws.cell(row=header_row, column=preview_col).fill = header_fill
    ws.cell(row=header_row, column=preview_col).alignment = center_align
    ws.column_dimensions[get_column_letter(preview_col)].width = 30

    for row_idx in range(header_row + 1, ws.max_row + 1):
        day_val = ws.cell(row=row_idx, column=1).value
        if day_val is None:
            continue
        for thumb_file in thumb_dir.glob(f"day{int(day_val):02d}_*.png"):
            try:
                img = XlImage(str(thumb_file))
                img.width = 180
                img.height = 180
                ws.add_image(img, f"{get_column_letter(preview_col)}{row_idx}")
                ws.row_dimensions[row_idx].height = 140
            except Exception:
                pass
            break

    wb.save(excel_path)
    wb.close()


# ═══════════════════════════════════════════════════════════════
# PHASE 4: Package everything into a zip
# ═══════════════════════════════════════════════════════════════

def package_output(output_dir: Path, callback: Callable) -> Path:
    """Create a downloadable zip of Excel + images."""
    callback("status", {"phase": "packaging", "message": "Creating download package..."})

    zip_path = output_dir / "MSRead_Content_Engine.zip"
    excel_path = output_dir / "MSRead_Content_Engine.xlsx"
    image_dir = output_dir / "images"

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        if excel_path.exists():
            zf.write(excel_path, "MSRead_Content_Engine.xlsx")
        if image_dir.exists():
            for img_file in sorted(image_dir.glob("*.png")):
                zf.write(img_file, f"images/{img_file.name}")

    size_mb = zip_path.stat().st_size / (1024 * 1024)
    callback("status", {
        "phase": "done",
        "message": f"Done! Package ready ({size_mb:.1f} MB)",
        "zip_size_mb": round(size_mb, 1),
    })
    return zip_path


# ═══════════════════════════════════════════════════════════════
# ORCHESTRATOR: Run the full pipeline
# ═══════════════════════════════════════════════════════════════

def run_pipeline(api_key: str, creative_brief: str, output_dir: Path, callback: Callable):
    """Run the full content engine pipeline."""
    try:
        client = genai.Client(api_key=api_key)

        start_date = datetime.now().strftime("%b %d, %Y")

        # Phase 1: Adapt content
        content = generate_adapted_content(client, creative_brief, start_date, callback)

        # Phase 2: Build Excel
        excel_path = build_excel(content, output_dir, callback)

        # Phase 3: Generate images
        generate_images(client, creative_brief, excel_path, output_dir, callback)

        # Phase 4: Package
        package_output(output_dir, callback)

    except Exception as e:
        callback("error", {"message": f"Pipeline failed: {str(e)}"})
        raise
