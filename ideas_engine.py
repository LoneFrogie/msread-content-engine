"""
MS. READ Content Engine — 30-Day Idea / Theme Planner

Generates a full month of content-calendar THEMES (the planning layer — not
the finished posts). The voice is lifestyle-led and soft-sell, inspired by how
H&M, J.Crew, Net-a-Porter and local MY brands (Hanya, Mis Claire, Machino,
Hani Mokhta) build brand through routines, moods and small everyday moments
rather than hard product pushes.

Output = structured JSON the frontend renders, plus an Excel planner that the
existing /api/download-excel endpoint serves.
"""

import json
import time
from pathlib import Path
from typing import Callable

from google import genai
from google.genai import types
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from sku_engine import _parse_json_robust, BRAND_NAME

# Default brand references (the art director's mix). Overridable from the UI.
DEFAULT_BRANDS = "J.Crew, Net-a-Porter, H&M, Hanya, Mis Claire, Machino, Hani Mokhta"

# Brand colours for the Excel planner
NAVY = "1A2238"
CORAL = "E8927C"
LIGHT = "F7F3F0"

IDEAS_PROMPT = """You are the social content strategist for {brand}, Malaysia's leading plus-size fashion brand (UK sizes 10-24+, founded 1997).

BRAND VOICE:
- Warm, empowering, inclusive, quietly confident.
- NEVER use diet-culture language or an apologetic tone about size.
- Malaysian English, premium but down-to-earth and relatable.
- Soft-sell: lead with routines, moods, plans and small real-life moments — the clothes fit naturally INTO the story; they are rarely the headline.

INSPIRATION BRANDS: {brands}
Study how these brands build desire through lifestyle and styling rather than promotion — "Running Late Go-To Outfit", "1 Blouse, 3 Ways", "Weekend Getaway Pieces", "Slow Mornings, Easy Dressing", "A Day Well Dressed". Borrow that *register* (calm, useful, aspirational), not their exact wording.

PLANNING BRIEF FROM THE ART DIRECTOR:
{theme_direction}

TASK:
Plan {num_days} days of content-calendar THEMES for {brand}'s {month} calendar. These are ideas/directions the team will later turn into posts — NOT the finished captions.

Target feel: about 60% lifestyle + styling inspiration, 25% wardrobe solutions, only 15% direct product. Make it specific to Malaysian life (heat, aircon, rain, long lunches, public holidays, weekend markets) where it fits.

Group the {num_days} days into weeks of 7 (Week 1 = days 1-7, Week 2 = days 8-14, and so on; the final week may be short). Give each week a short sub-theme label.

For EACH day provide:
- "day": the day number (integer)
- "title": the theme name (e.g. "Running Late Go-To Outfit") — punchy, 2-6 words
- "subtitle": a one-line description (e.g. "The formula that never fails.")
- "format": ONE suggested post format from: Carousel, Reel, Single Image, Story, Guide
- "hook": a short opening caption line in {brand}'s voice (max ~12 words) the team could actually use

Return a JSON object with these EXACT keys:
{{
  "month": "{month}",
  "theme_name": "<an overarching monthly theme, e.g. 'Everyday Moments, Thoughtfully Dressed'>",
  "content_mix": "<one short line, e.g. '60% lifestyle & styling · 25% wardrobe solutions · 15% product'>",
  "weeks": [
    {{
      "week_label": "Week 1: <weekly sub-theme>",
      "days": [
        {{ "day": 1, "title": "...", "subtitle": "...", "format": "...", "hook": "..." }}
      ]
    }}
  ],
  "bonus_themes": [
    {{ "category": "<group name, e.g. 'Community & Engagement'>", "items": ["<theme>", "<theme>", "..."] }}
  ]
}}

For "bonus_themes", give 4-6 groups of swap-in ideas (e.g. H&M-style, J.Crew-style, Local-brand-style, Net-a-Porter-style, Community & Engagement, Malaysian Lifestyle), each with 4-6 short theme titles, that can replace any day around launches or promos.

Respond with ONLY valid JSON, no markdown code fences."""


def generate_ideas(client, theme_direction: str, brands: str, month: str,
                   num_days: int, callback: Callable) -> dict:
    """Call Gemini to produce the structured month-of-themes JSON."""
    callback("status", {"phase": "generating_ideas",
                        "message": f"Planning {num_days} days of {month} content themes..."})

    prompt_text = IDEAS_PROMPT.format(
        brand=BRAND_NAME,
        brands=brands.strip() or DEFAULT_BRANDS,
        theme_direction=theme_direction.strip() or "(no extra direction — use your best judgement for a plus-size MY audience)",
        month=month.strip() or "the upcoming month",
        num_days=num_days,
    )

    last_error = None
    attempts = [(0.85, 5), (0.7, 15), (0.5, 30), (0.4, 45)]
    for attempt, (temp, wait) in enumerate(attempts):
        try:
            response = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=[types.Part.from_text(text=prompt_text)],
                config=types.GenerateContentConfig(
                    temperature=temp,
                    max_output_tokens=8000,
                    response_mime_type="application/json",
                ),
            )

            text = ""
            for part in response.candidates[0].content.parts:
                if part.text:
                    text += part.text

            text = text.strip()
            if text.startswith("```"):
                text = text.split("\n", 1)[1]
            if text.endswith("```"):
                text = text.rsplit("```", 1)[0]
            text = text.strip()

            content = _parse_json_robust(text)
            callback("status", {"phase": "ideas_generated",
                                "message": f"{num_days}-day theme plan ready"})
            return content

        except (json.JSONDecodeError, ValueError) as e:
            last_error = e
            if attempt < len(attempts) - 1:
                callback("status", {"phase": "generating_ideas",
                                    "message": "Retrying with adjusted parameters..."})
                time.sleep(wait)
        except Exception as e:
            last_error = e
            err_str = str(e)
            if ("503" in err_str or "UNAVAILABLE" in err_str or "overloaded" in err_str.lower()) and attempt < len(attempts) - 1:
                callback("status", {"phase": "generating_ideas",
                                    "message": f"Model busy, retrying in {wait}s..."})
                time.sleep(wait)
            else:
                raise

    raise last_error


def build_ideas_excel(content: dict, output_dir: Path, callback: Callable) -> Path:
    """Build a 2-sheet Excel planner (Calendar + Bonus Themes)."""
    callback("status", {"phase": "building_excel", "message": "Building Excel planner..."})

    month = content.get("month", "Content")
    wb = Workbook()

    navy_fill = PatternFill("solid", fgColor=NAVY)
    coral_fill = PatternFill("solid", fgColor=CORAL)
    light_fill = PatternFill("solid", fgColor=LIGHT)
    white_bold = Font(bold=True, color="FFFFFF", size=12)
    navy_bold = Font(bold=True, color=NAVY, size=11)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Sheet 1: Calendar ──
    ws = wb.active
    ws.title = "Calendar"
    ws["A1"] = f"{BRAND_NAME} — {month} Content Calendar"
    ws["A1"].font = Font(bold=True, color=NAVY, size=16)
    ws["A2"] = content.get("theme_name", "")
    ws["A2"].font = Font(italic=True, color=NAVY, size=12)
    ws["A3"] = content.get("content_mix", "")
    ws["A3"].font = Font(color="666666", size=10)

    headers = ["Day", "Theme", "One-liner", "Format", "Caption Hook"]
    hrow = 5
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hrow, column=c, value=h)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = wrap
        cell.border = border

    r = hrow + 1
    for week in content.get("weeks", []):
        # Week banner row
        ws.cell(row=r, column=1, value=week.get("week_label", "")).font = navy_bold
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = light_fill
            ws.cell(row=r, column=c).border = border
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
        for d in week.get("days", []):
            row_vals = [d.get("day", ""), d.get("title", ""), d.get("subtitle", ""),
                        d.get("format", ""), d.get("hook", "")]
            for c, v in enumerate(row_vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.alignment = wrap
                cell.border = border
            r += 1

    widths = [6, 30, 40, 14, 45]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

    # ── Sheet 2: Bonus Themes ──
    ws2 = wb.create_sheet("Bonus Themes")
    ws2["A1"] = "Swap-in Themes (replace any day around launches / promos)"
    ws2["A1"].font = Font(bold=True, color=NAVY, size=14)
    for c, h in enumerate(["Category", "Theme"], 1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.fill = coral_fill
        cell.font = white_bold
        cell.border = border
    r2 = 4
    for grp in content.get("bonus_themes", []):
        cat = grp.get("category", "")
        for item in grp.get("items", []):
            ws2.cell(row=r2, column=1, value=cat).border = border
            ws2.cell(row=r2, column=2, value=item).border = border
            r2 += 1
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 50

    safe_month = "".join(ch for ch in month if ch.isalnum() or ch in " -_").strip().replace(" ", "_") or "Content"
    xlsx_path = output_dir / f"MSRead_{safe_month}_Theme_Calendar.xlsx"
    wb.save(xlsx_path)
    return xlsx_path


# ═══════════════════════════════════════════════════════════════
# ORCHESTRATOR
# ═══════════════════════════════════════════════════════════════

def run_ideas_pipeline(api_key: str, theme_direction: str, brands: str, month: str,
                       output_dir: Path, callback: Callable, num_days: int = 30):
    """Run the idea/theme planning pipeline."""
    try:
        client = genai.Client(api_key=api_key)

        content = generate_ideas(client, theme_direction, brands, month, num_days, callback)

        # Save JSON for the frontend to render
        with open(output_dir / "ideas_content.json", "w") as f:
            json.dump(content, f, indent=2)

        # Build the downloadable Excel planner (best-effort)
        try:
            build_ideas_excel(content, output_dir, callback)
        except Exception as e:
            callback("status", {"phase": "building_excel",
                                "message": f"Excel skipped: {str(e)[:80]}"})

        callback("status", {
            "phase": "done",
            "message": "Theme calendar ready!",
            "ideas_content": content,
        })

    except Exception as e:
        callback("error", {"message": f"Pipeline failed: {str(e)}"})
        raise
